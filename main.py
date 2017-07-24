from bs4 import BeautifulSoup
import requests
import openpyxl

def getemail( cnpj ):
   empresa = "https://www.qualempresa.com.br/empresa/" + cnpj + "/"
   page = requests.get(empresa)
   soup = BeautifulSoup(page.content, 'html.parser')
   resultato_html = soup.find_all(id="main")
   resultato_html_str = str(resultato_html[0])
   strv = resultato_html_str.split('<li>E-mail:  <strong>', 1000000)
   email = strv[1].split('</strong></li>', 1)
   return email[0]

def gettelefone( cnpj ):
   empresa = "https://www.qualempresa.com.br/empresa/" + cnpj + "/"
   page = requests.get(empresa)
   soup = BeautifulSoup(page.content, 'html.parser')
   resultato_html = soup.find_all(id="main")
   resultato_html_str = str(resultato_html[0])
   strv = resultato_html_str.split('Telefone:  <strong>', 1000000)
   telefone = strv[1].split('</strong></li>', 1)
   return telefone[0]

def getcnpj( nome_restaurante ):
   restaurante = nome_restaurante
   restaurante = restaurante.replace(" ", "+")
   # busca nome empresarial
   busca = "http://www.cnpj.ninja/buscar/?keyword=" + restaurante + "&s=1"
   page = requests.get(busca)
   soup = BeautifulSoup(page.content, 'html.parser')
   resultato_html = soup.find_all(id="search_results")
   if not resultato_html:
      return "27691471000140"
   resultato_html_str = str(resultato_html[0])
   strv = resultato_html_str.split('<h3><a href="/', 1000000)
   link = strv[1].split('" title=', 1000000)
   link = "http://www.cnpj.ninja/" + link[0]
   # busca do cnpj
   pagina = requests.get(link)
   soupe = BeautifulSoup(pagina.content, 'html.parser')
   html_resultado = soupe.find_all(id="details")
   if not html_resultado:
       return "27691471000140"
   html_resultado_str = str(html_resultado[0])
   str_v = html_resultado_str.split('ero do CNPJ</h5><p>', 1000000)
   cnpj = str_v[1].split('</p><div class="adg"', 1000000)
   cnpj = cnpj[0]
   cnpj = cnpj.replace(".", "")
   cnpj = cnpj.replace("/", "")
   cnpj = cnpj.replace("-", "")
   return cnpj

wb = openpyxl.Workbook()
wb = openpyxl.load_workbook(filename = 'clientes.xlsx')
sheets = wb.sheetnames
ws = wb[sheets[0]]

for x in range(3, 14180):
   nome_empresarial = ws['A'+str(x)].value
   nome_empresarial = str(nome_empresarial)
   cnpj = getcnpj(nome_empresarial)
   telefone = gettelefone(cnpj)
   email = getemail(cnpj)
   print "nome: "+nome_empresarial+" | email: "+email+" | telefone: "+telefone




